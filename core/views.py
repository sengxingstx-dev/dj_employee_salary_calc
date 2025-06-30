import calendar
from datetime import time
from decimal import Decimal

import openpyxl
import pytz
from dateutil.relativedelta import relativedelta  # Added for month iteration
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse, QueryDict
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl.styles import Alignment, Font

from accounts.models import Account, Employee
from core.models import (
    Attendance,
    Bonuses,
    Deductions,
    SalaryCalculations,
    SalaryStructure,
)

from .forms import (
    BonusesForm,
    DeductionsForm,
    EmployeeForm,
    SalaryCalculationsForm,
    SalaryStructureForm,
)


def custom_404_view(request, exception):
    return render(request, "404.html", status=404)


@login_required
def home(request):
    employee = None

    if request.user.is_authenticated:
        user = request.user
        employee = get_object_or_404(Employee, user=user)

    context = {"employee": employee}
    return render(request, "core/clients/pages/home.html", context)


def about(request):
    return render(request, "core/clients/pages/about.html")


def contact(request):
    return render(request, "core/clients/pages/contact.html")


@login_required
def scan_in(request):
    if request.method == "POST":
        user = request.user
        employee = get_object_or_404(Employee, user=user)

        # Set the timezone to Asia/Vientiane
        tz = pytz.timezone("Asia/Vientiane")
        now = timezone.now().astimezone(tz)
        today = now.date()
        current_time = now.time()

        # Define shift times
        morning_start = time(8, 0)
        afternoon_start = time(12, 0)
        night_start = time(20, 0)
        work_end = time(16, 0)
        # day_end = time(23, 59, 59)

        # Determine shift
        if morning_start <= current_time < afternoon_start:
            shift = "Morning"
        elif afternoon_start <= current_time < night_start:
            shift = "Afternoon"
        else:
            shift = "Night"

        # Check if it's past work hours
        is_overtime = current_time > work_end

        # Create or get attendance record
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            date=today,
            defaults={
                "shift": shift,
                "hours_worked": 0,
                "overtime_hours": 0,
                "is_present": True,
                "scan_in_time": now,
            },
        )

        if not created:
            if attendance.scan_in_time:
                messages.warning(request, "You have already scanned in today.")
            else:
                attendance.scan_in_time = now
                attendance.shift = shift
                attendance.is_present = True
                attendance.save()
                messages.success(request, f"Scan in successful. Shift: {shift}")
        else:
            messages.success(request, f"Scan in successful. Shift: {shift}")

        # If scanning in after work hours, log overtime
        if is_overtime:
            # attendance.overtime_hours = (
            #     now - now.replace(hour=16, minute=0, second=0, microsecond=0)
            # ).total_seconds() / 3600
            # attendance.save()
            messages.info(
                request, "You are scanning in after regular work hours. Overtime will be logged."
            )

        return redirect("home")  # Redirect to an attendance list view

    # return render(request, "core/scan_in.html")
    return render(request, "core/clients/pages/home.html")


@login_required
def scan_out(request):
    if request.method == "POST":
        user = request.user
        employee = get_object_or_404(Employee, user=user)

        # Set the timezone to Asia/Vientiane
        tz = pytz.timezone("Asia/Vientiane")
        now = timezone.now().astimezone(tz)
        today = now.date()

        try:
            attendance = Attendance.objects.get(employee=employee, date=today)

            if not attendance.scan_in_time:
                messages.error(request, "You need to scan in first.")
            elif attendance.scan_out_time:
                messages.warning(request, "You have already scanned out today.")
            else:
                attendance.scan_out_time = now
                attendance.save()
                attendance.calculate_hours_worked()
                messages.success(request, "Scan out successful.")

                # Check if scanning out after work hours and log additional overtime
                work_end = now.replace(hour=16, minute=0, second=0, microsecond=0)
                if now > work_end:
                    # additional_overtime = (now - work_end).total_seconds() / 3600
                    # attendance.overtime_hours += additional_overtime
                    # attendance.save()
                    additional_overtime_float = (now - work_end).total_seconds() / 3600
                    additional_overtime_decimal = Decimal(str(additional_overtime_float))

                    # Ensure overtime_hours is not None if it could be (though model default is 0)
                    if attendance.overtime_hours is None:
                        attendance.overtime_hours = Decimal("0.00")
                    attendance.overtime_hours += additional_overtime_decimal
                    attendance.save(update_fields=["overtime_hours"])  # Save only the changed field

                    messages.info(
                        request,
                        f"Additional overtime logged: {additional_overtime_decimal:.2f} hours",
                    )

        except Attendance.DoesNotExist:
            messages.error(request, "No attendance record found for today. Please scan in first.")

        # return redirect("attendance_list")  # Redirect to an attendance list view
        return redirect("home")

    # return render(request, "core/scan_out.html")
    return render(request, "core/clients/pages/home.html")


# NOTE: Dashboard section
@login_required
def dashboard(request):
    total_employees = Employee.objects.count()
    active_employees = Employee.objects.filter(status__iexact="active").count()
    inactive_employees = total_employees - active_employees

    current_month = timezone.now().month
    current_year = timezone.now().year

    salaries_current_month = SalaryCalculations.objects.filter(
        month_year__year=current_year, month_year__month=current_month
    )
    salaries_paid_current_month = salaries_current_month.filter(status="PAID").count()
    salaries_pending_current_month = salaries_current_month.filter(status="PENDING").count()
    total_salaries_processed_current_month = salaries_current_month.count()

    # Optional: Recent activity (e.g., last 5 salary calculations)
    recent_salary_calculations = SalaryCalculations.objects.select_related("employee").order_by(
        "-generated_at"
    )[:5]

    # Data for Salary Overview Chart (last 6 months)
    chart_labels = []
    chart_data_paid = []
    chart_data_pending = []

    for i in range(5, -1, -1):  # Last 6 months, including current
        month_date = timezone.now().date().replace(day=1) - relativedelta(months=i)
        chart_labels.append(month_date.strftime("%b %Y"))  # e.g., Jul 2024

        salaries_in_month = SalaryCalculations.objects.filter(
            month_year__year=month_date.year, month_year__month=month_date.month
        )

        total_paid_in_month = (
            salaries_in_month.filter(status="PAID").aggregate(total=Sum("net_salary"))["total"] or 0
        )
        chart_data_paid.append(float(total_paid_in_month))  # Chart.js expects numbers

        total_pending_in_month = (
            salaries_in_month.filter(status="PENDING").aggregate(total=Sum("net_salary"))["total"]
            or 0
        )
        chart_data_pending.append(float(total_pending_in_month))

    print("Chart Labels:", chart_labels)
    print("Chart Data Paid:", chart_data_paid)
    print("Chart Data Pending:", chart_data_pending)

    context = {
        "total_employees": total_employees,
        "active_employees": active_employees,
        "inactive_employees": inactive_employees,
        "salaries_paid_current_month": salaries_paid_current_month,
        "salaries_pending_current_month": salaries_pending_current_month,
        "total_salaries_processed_current_month": total_salaries_processed_current_month,
        "recent_salary_calculations": recent_salary_calculations,
        "chart_labels": chart_labels,
        "chart_data_paid": chart_data_paid,
        "chart_data_pending": chart_data_pending,
    }
    return render(request, "core/dashboard/pages/dashboard.html", context)


def manage_profile(request):
    return render(request, "core/dashboard/pages/manage-profile.html")


@login_required
def manage_users(request):
    query = request.GET.get("search", "")
    users = Account.objects.all().order_by("-date_joined")

    if query:
        # Search across multiple fields
        users = users.filter(Q(email__icontains=query))

    msg = "Are you sure you want to delete this user?"
    context = {
        "users": users,
        "delete_confirm_msg": msg,
    }
    return render(request, "core/dashboard/pages/manage-users.html", context)


@login_required
def delete_user(request, pk):
    user = Account.objects.get(pk=pk)

    if request.method == "POST":
        user.delete()
        return redirect("manage-users")

    return render(request, "core/dashboard/pages/manage-users.html")


@login_required
def manage_employees(request):
    # employee = Employee.objects.get(user=request.user)
    query = request.GET.get("search", "")
    employees = Employee.objects.all().order_by("-updated_at")

    if query:
        # Search across multiple fields
        employees = employees.filter(
            Q(name__icontains=query)
            | Q(contact_number__icontains=query)
            | Q(user__email__icontains=query)
        )

    if request.method == "POST":
        form = EmployeeForm(request.POST)
        email = request.POST.get("email")
        password = request.POST.get("password")

        if form.is_valid():
            hashed_password = make_password(password)
            user = Account.objects.create(email=email, password=hashed_password)
            employee = form.save(commit=True)
            employee.user = user
            employee.save()
            user.save()
            messages.success(request, "Employee created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        # "employee": employee,
        "employees": employees,
        "delete_confirm_msg": "Are you sure you want to delete this employee?",
    }
    return render(request, "core/dashboard/pages/manage-employees.html", context)


@login_required
def edit_employee(request, pk):
    # Get the employee to update
    employee = Employee.objects.get(id=pk)
    data = QueryDict(request.body)  # Use QueryDict to parse the request body
    emp_form = EmployeeForm(data, instance=employee)  # Bind the data to the existing instance

    if request.method == "POST":
        if emp_form.is_valid():
            emp_form.save()
            messages.success(request, "Employee updated successfully.")
        else:
            return JsonResponse({"error": emp_form.errors}, status=400)
        return redirect("manage-employees")
    else:
        emp_form = EmployeeForm(data, instance=employee)

    return render(request, "core/dashboard/pages/manage-employees.html")


@login_required
def delete_employee(request, pk):
    employee = Employee.objects.get(pk=pk)

    if request.method == "POST":
        employee.delete()
        return redirect("manage-employees")

    return render(request, "core/dashboard/pages/manage-employees.html")


@login_required
def manage_salary_structures(request):
    query = request.GET.get("search", "")
    employees = Employee.objects.all().order_by("-updated_at")
    structures = SalaryStructure.objects.all().order_by("-updated_at")

    if query:
        # Search across multiple fields
        structures = structures.filter(
            Q(basic_salary__icontains=query) | Q(bonus_percentage__icontains=query)
        )

    if request.method == "POST":
        form = SalaryStructureForm(request.POST)

        if form.is_valid():
            structure = form.save(commit=True)
            structure.save()
            messages.success(request, "Salary Structure created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "employees": employees,
        "structures": structures,
        "delete_confirm_msg": "Are you sure you want to delete this salary structure?",
    }
    return render(request, "core/dashboard/pages/manage-salary-structures.html", context)


@login_required
def edit_salary_structure(request, pk):
    structure = SalaryStructure.objects.get(id=pk)
    data = QueryDict(request.body)
    structure_form = SalaryStructureForm(data, instance=structure)

    if request.method == "POST":
        if structure_form.is_valid():
            structure_form.save()
            messages.success(request, "Salary structure updated successfully.")
        else:
            return JsonResponse({"error": structure_form.errors}, status=400)
        return redirect("manage-salary-structures")
    else:
        structure_form = SalaryStructureForm(data, instance=structure)

    return render(request, "core/dashboard/pages/manage-salary-structures.html")


@login_required
def delete_salary_structure(request, pk):
    structure = SalaryStructure.objects.get(pk=pk)

    if request.method == "POST":
        structure.delete()
        return redirect("manage-salary-structures")

    return render(request, "core/dashboard/pages/manage-salary-structures.html")


@login_required
def manage_deductions(request):
    query = request.GET.get("search", "")
    employees = Employee.objects.all().order_by("-updated_at")
    deductions = Deductions.objects.all().order_by("-updated_at")

    if query:
        # Search across multiple fields
        deductions = deductions.filter(Q(date__icontains=query) | Q(amount__icontains=query))

    if request.method == "POST":
        form = DeductionsForm(request.POST)

        if form.is_valid():
            deduction = form.save(commit=True)
            deduction.save()
            messages.success(request, "Deduction created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "employees": employees,
        "deductions": deductions,
        "delete_confirm_msg": "Are you sure you want to delete this deduction?",
    }
    return render(request, "core/dashboard/pages/manage-deductions.html", context)


@login_required
def edit_deduction(request, pk):
    deduction = Deductions.objects.get(id=pk)
    data = QueryDict(request.body)
    deduction_form = DeductionsForm(data, instance=deduction)

    if request.method == "POST":
        if deduction_form.is_valid():
            deduction_form.save()
            messages.success(request, "Deduction updated successfully.")
        else:
            return JsonResponse({"error": deduction_form.errors}, status=400)
        return redirect("manage-deductions")
    else:
        deduction_form = DeductionsForm(data, instance=deduction)

    return render(request, "core/dashboard/pages/manage-deductions.html")


@login_required
def delete_deduction(request, pk):
    deduction = Deductions.objects.get(pk=pk)

    if request.method == "POST":
        deduction.delete()
        return redirect("manage-deductions")

    return render(request, "core/dashboard/pages/manage-deductions.html")


@login_required
def manage_bonuses(request):
    query = request.GET.get("search", "")
    employees = Employee.objects.all().order_by("-updated_at")
    bonuses = Bonuses.objects.all().order_by("-updated_at")

    if query:
        # Search across multiple fields
        bonuses = bonuses.filter(Q(date__icontains=query) | Q(amount__icontains=query))

    if request.method == "POST":
        form = BonusesForm(request.POST)

        if form.is_valid():
            bonus = form.save(commit=True)
            bonus.save()
            messages.success(request, "Deduction created successfully.")
        else:
            print("Form is not valid: ", form.errors)

    context = {
        "employees": employees,
        "bonuses": bonuses,
        "delete_confirm_msg": "Are you sure you want to delete this bonus?",
    }
    return render(request, "core/dashboard/pages/manage-bonuses.html", context)


@login_required
def edit_bonus(request, pk):
    bonus = Bonuses.objects.get(id=pk)
    data = QueryDict(request.body)
    bonus_form = BonusesForm(data, instance=bonus)

    if request.method == "POST":
        if bonus_form.is_valid():
            bonus_form.save()
            messages.success(request, "Bonus updated successfully.")
        else:
            return JsonResponse({"error": bonus_form.errors}, status=400)
        return redirect("manage-bonuses")
    else:
        bonus_form = BonusesForm(data, instance=bonus)

    return render(request, "core/dashboard/pages/manage-bonuses.html")


@login_required
def delete_bonus(request, pk):
    bonus = Bonuses.objects.get(pk=pk)

    if request.method == "POST":
        bonus.delete()
        return redirect("manage-bonuses")

    return render(request, "core/dashboard/pages/manage-bonuses.html")


def manage_products(request):
    context = {
        "delete_confirm_msg": "Are you sure you want to delete this product?",
    }
    return render(request, "core/dashboard/pages/manage-products.html", context)


@login_required
def manage_attendance(request):
    query = request.GET.get("search", "")
    # Fetch attendance records, ordering by date descending, then employee name
    # Use select_related to efficiently fetch related employee data
    attendance_list = (
        Attendance.objects.select_related("employee").all().order_by("-date", "employee__name")
    )

    if query:
        # Search across multiple fields: employee name, date (exact match or partial), shift
        attendance_list = attendance_list.filter(
            Q(employee__name__icontains=query)
            | Q(date__icontains=query)  # Allows searching like '2024-07' or '2024-07-26'
            | Q(shift__icontains=query)
        )

    # Pagination (Optional but recommended for large datasets)
    page = request.GET.get("page", 1)
    paginator = Paginator(attendance_list, 10)  # Show 10 records per page
    try:
        attendance_records = paginator.page(page)
    except PageNotAnInteger:
        attendance_records = paginator.page(1)
    except EmptyPage:
        attendance_records = paginator.page(paginator.num_pages)

    context = {
        "attendance_records": attendance_records,
        "search_query": query,  # Pass the query back for display in the search box
        "delete_confirm_msg": "Are you sure you want to delete this attendance record?",
    }
    return render(request, "core/dashboard/pages/manage-attendance.html", context)


# --------- Salary Calculations Section ---------
# NOTE: Salary Calculations Section
@login_required
def manage_salary_calculations(request):
    current_year = timezone.now().year
    years = range(current_year - 5, current_year + 2)  # Range of years for dropdown
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]

    selected_year = request.GET.get("year", str(current_year))
    selected_month = request.GET.get("month", str(timezone.now().month))
    employee_filter = request.GET.get("employee", "")
    status_filter = request.GET.get("status", "")

    calculations_list = SalaryCalculations.objects.select_related("employee").all()

    if selected_year and selected_month:
        try:
            year_int = int(selected_year)
            month_int = int(selected_month)
            calculations_list = calculations_list.filter(
                month_year__year=year_int, month_year__month=month_int
            )
        except ValueError:
            messages.error(request, "Invalid year or month selected.")

    if employee_filter:
        calculations_list = calculations_list.filter(employee_id=employee_filter)

    if status_filter:
        calculations_list = calculations_list.filter(status=status_filter)

    # Handle Salary Generation
    if request.method == "POST" and "generate_salary" in request.POST:
        gen_year = request.POST.get("generate_year")
        gen_month = request.POST.get("generate_month")
        gen_employee_id = request.POST.get(
            "generate_employee_id", None
        )  # Optional: for single employee

        if not gen_year or not gen_month:
            messages.error(request, "Year and Month are required for salary generation.")
        else:
            try:
                year = int(gen_year)
                month = int(gen_month)
                target_month_date = timezone.datetime(year, month, 1).date()

                employees_to_process = Employee.objects.filter(
                    status="active"
                )  # Or your criteria for active employees
                if gen_employee_id:
                    employees_to_process = employees_to_process.filter(id=gen_employee_id)

                generated_count = 0
                failed_count = 0

                for emp in employees_to_process:
                    try:
                        salary_structure = SalaryStructure.objects.get(
                            employee=emp
                        )  # Assuming one active structure

                        # Aggregate Attendance
                        attendance_for_month = Attendance.objects.filter(
                            employee=emp, date__year=year, date__month=month
                        )
                        total_hours = (
                            attendance_for_month.aggregate(Sum("hours_worked"))["hours_worked__sum"]
                            or 0
                        )
                        total_overtime = (
                            attendance_for_month.aggregate(Sum("overtime_hours"))[
                                "overtime_hours__sum"
                            ]
                            or 0
                        )

                        # Aggregate Deductions
                        total_deductions = (
                            Deductions.objects.filter(
                                employee=emp, date__year=year, date__month=month
                            ).aggregate(Sum("amount"))["amount__sum"]
                            or 0
                        )

                        # Aggregate Bonuses
                        total_bonuses = (
                            Bonuses.objects.filter(
                                employee=emp, date__year=year, date__month=month
                            ).aggregate(Sum("amount"))["amount__sum"]
                            or 0
                        )

                        # Calculate Gross and Net Salary
                        # This is a basic calculation, adjust as per your rules
                        # E.g., if basic_salary is monthly fixed:
                        gross_salary = (
                            salary_structure.basic_salary
                            + (total_overtime * salary_structure.overtime_rate)
                            + total_bonuses
                        )
                        net_salary = gross_salary - total_deductions

                        # Create or Update Salary Calculation
                        calculation, created = SalaryCalculations.objects.update_or_create(
                            employee=emp,
                            month_year=target_month_date,
                            defaults={
                                "basic_salary_snapshot": salary_structure.basic_salary,
                                "overtime_rate_snapshot": salary_structure.overtime_rate,
                                "total_hours_worked": total_hours,
                                "total_overtime_hours": total_overtime,
                                "total_deductions_amount": total_deductions,
                                "total_bonuses_amount": total_bonuses,
                                "gross_salary": gross_salary,
                                "net_salary": net_salary,
                                "status": "PENDING",  # Reset to pending on re-generation if not paid
                                "generated_at": timezone.now(),
                            },
                        )
                        if not created and calculation.status == "PAID":
                            # If overwriting a paid salary, perhaps set to pending or add a note.
                            # For now, it will update if defaults change.
                            messages.warning(
                                request,
                                f"Re-generated PENDING salary for {emp.name} for {target_month_date.strftime('%B %Y')}. Previous status was PAID.",
                            )
                            calculation.status = "PENDING"  # Ensure it's pending
                            calculation.paid_at = None
                            calculation.payment_method = None
                            calculation.save()

                        generated_count += 1
                    except SalaryStructure.DoesNotExist:
                        messages.error(
                            request, f"No salary structure found for {emp.name}. Skipped."
                        )
                        failed_count += 1
                    except Exception as e:
                        messages.error(request, f"Error generating salary for {emp.name}: {str(e)}")
                        failed_count += 1

                if generated_count > 0:
                    messages.success(
                        request, f"Successfully generated/updated {generated_count} salary records."
                    )
                if failed_count > 0:
                    messages.error(
                        request, f"Failed to generate/update {failed_count} salary records."
                    )
                return redirect("manage-salary-calculations")  # Redirect to refresh list

            except ValueError:
                messages.error(request, "Invalid year or month for generation.")
            except Exception as e:
                messages.error(request, f"An unexpected error occurred during generation: {str(e)}")

    paginator = Paginator(
        calculations_list.order_by("-month_year", "employee__name"), 10
    )  # 10 per page
    page = request.GET.get("page")
    try:
        calculations = paginator.page(page)
    except PageNotAnInteger:
        calculations = paginator.page(1)
    except EmptyPage:
        calculations = paginator.page(paginator.num_pages)

    all_employees = Employee.objects.all().order_by("name")
    status_choices = SalaryCalculations.STATUS_CHOICES

    context = {
        "salary_calculations": calculations,
        "employees": all_employees,  # For filter dropdown
        "years": years,
        "months": months,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "selected_employee_id": employee_filter,
        "selected_status": status_filter,
        "status_choices": status_choices,
        "delete_confirm_msg": "Are you sure you want to delete this salary calculation record?",
    }
    return render(request, "core/dashboard/pages/manage-salary-calculations.html", context)


@login_required
def view_salary_calculation(request, pk):
    calculation = get_object_or_404(SalaryCalculations, pk=pk)
    context = {
        "calculation": calculation,
    }
    return render(request, "core/dashboard/pages/view-salary-calculation.html", context)


@login_required
def edit_salary_calculation(request, pk):
    calculation = get_object_or_404(SalaryCalculations, pk=pk)
    if request.method == "POST":
        form = SalaryCalculationsForm(request.POST, instance=calculation)
        if form.is_valid():
            updated_calc = form.save(commit=False)
            if updated_calc.status == "PAID" and not updated_calc.paid_at:
                updated_calc.paid_at = timezone.now()  # Auto-set paid_at if marking as PAID
            elif updated_calc.status != "PAID":
                updated_calc.paid_at = None  # Clear paid_at if not PAID
                updated_calc.payment_method = None  # Clear payment_method if not PAID

            updated_calc.save()
            messages.success(request, "Salary calculation updated successfully.")
            return redirect("manage-salary-calculations")
        else:
            messages.error(request, f"Error updating salary calculation: {form.errors}")
    else:
        form = SalaryCalculationsForm(instance=calculation)

    context = {
        "form": form,
        "calculation": calculation,
        "page_title": f"ແກ້ໄຂເງິນເດືອນ: {calculation.employee.name} - {calculation.month_year.strftime('%B %Y')}",
    }
    # This view would typically render its own template or a modal within the list page.
    # For simplicity, let's assume it redirects back or uses a dedicated edit page.
    # If using a modal, you'd return JsonResponse for AJAX.
    # For now, let's create a simple edit page template (or reuse parts of manage-employees for form rendering)
    return render(request, "core/dashboard/pages/edit-salary-calculation.html", context)


@login_required
def delete_salary_calculation(request, pk):
    calculation = get_object_or_404(SalaryCalculations, pk=pk)
    if request.method == "POST":
        calculation.delete()
        messages.success(request, "Salary calculation deleted successfully.")
        return redirect("manage-salary-calculations")
    # Typically, deletion is confirmed via a modal on the list page,
    # or a separate confirmation page.
    # This direct POST delete is functional but might lack user confirmation step in UI.
    return redirect("manage-salary-calculations")  # Or render a confirmation page


# --------- Export Section ---------
# NOTE: Export to excel section
@login_required
def export_employees_to_excel(request):
    """Exports employee data to an Excel file."""

    # Define the response object for the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="employees_report.xlsx"'

    # Create an Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"

    # Define the header row
    columns = [
        "No.",
        "Name",
        "Email",
        "Position",
        "Department",
        "Contact Number",
        "Bank Account Number",
        "Employment Date",
        "Status",
    ]
    row_num = 1

    # Write the header row to the worksheet
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        # Optional: Add some basic styling to the header
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    # Get employee data (consider applying search filter if needed, similar to manage_employees)
    # For simplicity, exporting all employees here.
    employees = (
        Employee.objects.select_related("user").all().order_by("name")
    )  # Use select_related for efficiency

    # Write data rows
    for i, emp in enumerate(employees, start=1):
        row_num += 1
        user_email = emp.user.email if emp.user else "N/A"
        user_status = "Online" if emp.user and emp.user.is_active else "Offline"

        row = [
            i,
            emp.name,
            user_email,
            emp.position,
            emp.department,
            emp.contact_number,
            emp.bank_account_num,
            (
                emp.employment_date.strftime("%Y-%m-%d") if emp.employment_date else "N/A"
            ),  # Format date
            user_status,
        ]

        # Write the row to the worksheet
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            # Optional: Align text columns to the left
            if isinstance(cell_value, str):
                cell.alignment = openpyxl.styles.Alignment(horizontal="left")

    # Optional: Adjust column widths
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:  # Adjust letters based on your columns
        worksheet.column_dimensions[col_letter].width = 20
    worksheet.column_dimensions["A"].width = 5  # For No.

    # Save the workbook to the HttpResponse
    workbook.save(response)

    return response


@login_required
def export_salary_structures_to_excel(request):
    """Exports salary structure data to an Excel file."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="salary_structures_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Salary Structures Report"

    columns = [
        "No.",
        "Employee Name",
        "Basic Salary (LAK)",
        "Overtime Rate (LAK)",
        "Bonus Percentage (%)",
        "Created At",
        "Updated At",
    ]
    row_num = 1

    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = center_alignment

    search_query = request.GET.get("search", "")
    structures_list = (
        SalaryStructure.objects.select_related("employee")
        .all()
        .order_by("-updated_at", "employee__name")
    )

    if search_query:
        structures_list = structures_list.filter(
            Q(employee__name__icontains=search_query)
            | Q(basic_salary__icontains=search_query)
            | Q(bonus_percentage__icontains=search_query)
        )

    for i, structure in enumerate(structures_list, start=1):
        row_num += 1
        row_data = [
            i,
            structure.employee.name,
            f"{structure.basic_salary:,.0f}",
            f"{structure.overtime_rate:,.0f}",
            f"{structure.bonus_percentage:.2f}",  # Assuming bonus_percentage is stored like 5.00 for 5%
            structure.created_at.astimezone(pytz.timezone("Asia/Vientiane")).strftime(
                "%Y-%m-%d %H:%M"
            ),
            structure.updated_at.astimezone(pytz.timezone("Asia/Vientiane")).strftime(
                "%Y-%m-%d %H:%M"
            ),
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if col_num in [3, 4, 5]:  # Amount/Percentage columns
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = left_alignment

    worksheet.column_dimensions["A"].width = 5
    worksheet.column_dimensions["B"].width = 25  # Employee Name
    worksheet.column_dimensions["C"].width = 20  # Basic Salary
    worksheet.column_dimensions["D"].width = 20  # Overtime Rate
    worksheet.column_dimensions["E"].width = 20  # Bonus Percentage
    worksheet.column_dimensions["F"].width = 20  # Created At
    worksheet.column_dimensions["G"].width = 20  # Updated At

    workbook.save(response)
    return response


@login_required
def export_deductions_to_excel(request):
    """Exports deductions data to an Excel file."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="deductions_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Deductions Report"

    columns = [
        "No.",
        "Employee Name",
        "Date",
        "Reason",
        "Amount (LAK)",
        "Created At",
        "Updated At",
    ]
    row_num = 1

    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = center_alignment

    search_query = request.GET.get("search", "")
    deductions_list = (
        Deductions.objects.select_related("employee").all().order_by("-date", "employee__name")
    )

    if search_query:
        deductions_list = deductions_list.filter(
            Q(employee__name__icontains=search_query)
            | Q(reason__icontains=search_query)
            | Q(
                date__icontains=search_query
            )  # Basic date search, might need more robust parsing for specific formats
        )

    for i, deduction in enumerate(deductions_list, start=1):
        row_num += 1
        row_data = [
            i,
            deduction.employee.name,
            deduction.date.strftime("%Y-%m-%d"),
            deduction.reason,
            f"{deduction.amount:,.0f}",  # Formatted amount
            deduction.created_at.astimezone(pytz.timezone("Asia/Vientiane")).strftime(
                "%Y-%m-%d %H:%M"
            ),
            deduction.updated_at.astimezone(pytz.timezone("Asia/Vientiane")).strftime(
                "%Y-%m-%d %H:%M"
            ),
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if col_num == 5:  # Amount column
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = left_alignment

    worksheet.column_dimensions["A"].width = 5
    worksheet.column_dimensions["B"].width = 25  # Employee Name
    worksheet.column_dimensions["C"].width = 12  # Date
    worksheet.column_dimensions["D"].width = 30  # Reason
    worksheet.column_dimensions["E"].width = 15  # Amount
    worksheet.column_dimensions["F"].width = 20  # Created At
    worksheet.column_dimensions["G"].width = 20  # Updated At

    workbook.save(response)
    return response


@login_required
def export_bonuses_to_excel(request):
    """Exports bonuses data to an Excel file."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="bonuses_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Bonuses Report"

    columns = ["No.", "Employee Name", "Date", "Reason", "Amount (LAK)", "Created At", "Updated At"]
    row_num = 1

    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = center_alignment

    search_query = request.GET.get("search", "")
    bonuses_list = (
        Bonuses.objects.select_related("employee").all().order_by("-date", "employee__name")
    )

    if search_query:
        bonuses_list = bonuses_list.filter(
            Q(employee__name__icontains=search_query)
            | Q(reason__icontains=search_query)
            | Q(date__icontains=search_query)
        )

    for i, bonus in enumerate(bonuses_list, start=1):
        row_num += 1
        row_data = [
            i,
            bonus.employee.name,
            bonus.date.strftime("%Y-%m-%d"),
            bonus.reason,
            f"{bonus.amount:,.0f}",
            bonus.created_at.astimezone(pytz.timezone("Asia/Vientiane")).strftime("%Y-%m-%d %H:%M"),
            bonus.updated_at.astimezone(pytz.timezone("Asia/Vientiane")).strftime("%Y-%m-%d %H:%M"),
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if col_num == 5:  # Amount column
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = left_alignment

    worksheet.column_dimensions["A"].width = 5
    worksheet.column_dimensions["B"].width = 25
    worksheet.column_dimensions["C"].width = 12
    worksheet.column_dimensions["D"].width = 30
    worksheet.column_dimensions["E"].width = 15
    worksheet.column_dimensions["F"].width = 20
    worksheet.column_dimensions["G"].width = 20

    workbook.save(response)
    return response


@login_required
def export_attendance_to_excel(request):
    """Exports attendance data to an Excel file."""

    # Define the response object for the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="attendance_report.xlsx"'

    # Create an Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Attendance Report"

    # Define the header row
    columns = [
        "No.",
        "Employee Name",
        "Date",
        "Shift",
        "Scan In Time",
        "Scan Out Time",
        "Hours Worked",
        "Overtime Hours",
        "Status",
    ]
    row_num = 1

    # Write the header row to the worksheet
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = center_alignment

    # Get attendance data (consider applying search filter if needed)
    # For simplicity, exporting all attendance here, ordered like the manage page.
    # You could optionally reuse the search logic from manage_attendance if needed.
    attendance_records = (
        Attendance.objects.select_related("employee").all().order_by("-date", "employee__name")
    )

    # Write data rows
    for i, att in enumerate(attendance_records, start=1):
        row_num += 1
        status = "Present" if att.is_present else "Absent"
        # Format times safely, handling None values
        scan_in_str = (
            att.scan_in_time.astimezone(pytz.timezone("Asia/Vientiane")).strftime("%H:%M:%S")
            if att.scan_in_time
            else "--"
        )
        scan_out_str = (
            att.scan_out_time.astimezone(pytz.timezone("Asia/Vientiane")).strftime("%H:%M:%S")
            if att.scan_out_time
            else "--"
        )

        row = [
            i,
            att.employee.name,
            att.date.strftime("%Y-%m-%d"),  # Format date
            att.get_shift_display(),
            scan_in_str,
            scan_out_str,
            f"{att.hours_worked:.2f}" if att.hours_worked is not None else "0.00",  # Format decimal
            (
                f"{att.overtime_hours:.2f}" if att.overtime_hours is not None else "0.00"
            ),  # Format decimal
            status,
        ]

        # Write the row to the worksheet
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            # Align text left, numbers/times potentially center or right if desired
            if isinstance(cell_value, (int, float)) or col_num in [
                5,
                6,
                7,
                8,
            ]:  # Align numbers/times center
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment

    # Optional: Adjust column widths
    worksheet.column_dimensions["A"].width = 5  # No.
    worksheet.column_dimensions["B"].width = 25  # Employee Name
    worksheet.column_dimensions["C"].width = 12  # Date
    worksheet.column_dimensions["D"].width = 15  # Shift
    worksheet.column_dimensions["E"].width = 15  # Scan In
    worksheet.column_dimensions["F"].width = 15  # Scan Out
    worksheet.column_dimensions["G"].width = 15  # Hours Worked
    worksheet.column_dimensions["H"].width = 15  # Overtime
    worksheet.column_dimensions["I"].width = 10  # Status

    # Save the workbook to the HttpResponse
    workbook.save(response)

    return response


@login_required
def export_salary_calculations_to_excel(request):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="salary_calculations_report.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Salary Calculations"

    columns = [
        "No.",
        "Employee",
        "Month-Year",
        "Basic Salary",
        "Overtime Rate",
        "Total Hours Worked",
        "Total Overtime Hours",
        "Total Deductions",
        "Total Bonuses",
        "Gross Salary",
        "Net Salary",
        "Status",
        "Payment Method",
        "Paid At",
        "Generated At",
        "Notes",
    ]
    row_num = 1

    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.alignment = center_alignment

    # Apply filters from request.GET if needed, similar to manage_salary_calculations view
    selected_year = request.GET.get("year")
    selected_month = request.GET.get("month")
    employee_filter = request.GET.get("employee")
    status_filter = request.GET.get("status")

    queryset = (
        SalaryCalculations.objects.select_related("employee")
        .all()
        .order_by("-month_year", "employee__name")
    )

    if selected_year and selected_month:
        try:
            queryset = queryset.filter(
                month_year__year=int(selected_year), month_year__month=int(selected_month)
            )
        except ValueError:
            pass  # Ignore invalid filter
    if employee_filter:
        queryset = queryset.filter(employee_id=employee_filter)
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    for i, calc in enumerate(queryset, start=1):
        row_num += 1
        paid_at_str = calc.paid_at.strftime("%Y-%m-%d %H:%M") if calc.paid_at else "--"
        generated_at_str = (
            calc.generated_at.strftime("%Y-%m-%d %H:%M") if calc.generated_at else "--"
        )

        row_data = [
            i,
            calc.employee.name,
            calc.month_year.strftime("%B %Y"),
            f"{calc.basic_salary_snapshot:.2f}",
            f"{calc.overtime_rate_snapshot:.2f}",
            f"{calc.total_hours_worked:.2f}",
            f"{calc.total_overtime_hours:.2f}",
            f"{calc.total_deductions_amount:.2f}",
            f"{calc.total_bonuses_amount:.2f}",
            f"{calc.gross_salary:.2f}",
            f"{calc.net_salary:.2f}",
            calc.get_status_display(),
            calc.get_payment_method_display() if calc.payment_method else "--",
            paid_at_str,
            generated_at_str,
            calc.notes,
        ]
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = left_alignment  # Default left

    # Adjust column widths (example)
    worksheet.column_dimensions["B"].width = 25  # Employee Name
    for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
        worksheet.column_dimensions[col_letter].width = 18
    worksheet.column_dimensions["A"].width = 5

    workbook.save(response)
    return response
